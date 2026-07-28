#!/usr/bin/env python3
"""THE mandatory queue gate. Every recipient list must pass through here.

Usage:
  ../.venv/bin/python build_queue.py <input.csv> <output.csv> \
      --campaign <key> [--region nyc|any] [--exclude-list other-queue.csv]... \
      [--exclude-all-campaigns]

Runs EVERY check, every time — no opt-outs:
  1. email syntax (strict: no consecutive dots, real TLD, sane local part)
  2. junk/test/fake addresses and test domains
  3. typo domains (gamail, gnail, .con, ...)
  4. role/service addresses (info@, guestlist.service@, ...) incl. separator
     suffixes; known non-person domains (eventbrite.com)
  5. suppressed.csv + Klaviyo Profile Exclusions + master Emailable=no
     (ALL THREE MANDATORY — missing input is a hard error, never a skip)
  6. sent-log-*.json subtraction, scoped to --campaign (a future campaign
     must NOT silently exclude everyone ever emailed; --exclude-all-campaigns
     restores the everyone-ever behavior when that is actually wanted)
  7. duplicate aggregation: rows are grouped per email FIRST, and any
     disqualifying state on ANY row (unsubscribed) disqualifies the person
  8. region rules (--region nyc): far states/cities dropped, full state
     names normalized, location-updates.csv moves WIN over stale source data
  9. cross-queue overlap: hard-fails if the output would overlap any other
     manifest-signed queue for the same campaign (parallel-channel safety)

Output: <output.csv> + <output.csv>.manifest.json + <output.csv>.dropped.csv.
The manifest binds sha256, campaign, region, output basename, gate version
and source-file hashes; the sender refuses lists whose manifest is missing,
mismatched, from another campaign, or renamed. Publishing is atomic and any
failed build INVALIDATES the previous manifest first, so a crashed build can
never leave a stale-but-valid queue behind (Codex audit 2026-07-28, #3-#5).

Apple private-relay addresses (privaterelay.appleid.com etc.) are NOT junk:
Apple documents them as unique per-user forwarders for real people. Only
bounce-driven suppression removes them (Codex audit #11).

GATE_TEST_ROOT (env) redirects the data root for the fixture test suite ONLY.
Production invocations must never set it.
"""
import argparse
import csv
import glob
import hashlib
import json
import os
import re
import sys
import time
from datetime import datetime, timezone
from pathlib import Path

GATE_VERSION = 2
HERE = Path(__file__).resolve().parent
DATA_ROOT = Path(os.environ["GATE_TEST_ROOT"]) if os.environ.get("GATE_TEST_ROOT") else HERE.parent

# Strict-ish syntax: printable local part without leading/trailing/double
# dots, hyphen-safe domain labels, alphabetic TLD >= 2. Stdlib on purpose —
# an email-validation dependency adds supply-chain surface to a PII folder
# for marginal gain at this list size; bounces remain the final validator.
EMAIL_RE = re.compile(
    r"^(?!\.)[A-Za-z0-9!#$%&'*+/=?^_`{|}~.-]+(?<!\.)@"
    r"[A-Za-z0-9]([A-Za-z0-9-]*[A-Za-z0-9])?(\.[A-Za-z0-9]([A-Za-z0-9-]*[A-Za-z0-9])?)*\.[A-Za-z]{2,}$"
)
JUNK_LOCAL = re.compile(r"^(test|testing|fake|asdf+|abc+|example|sample|dummy|noreply|no-reply|donotreply)\d*$", re.I)
JUNK_DOMAIN = re.compile(r"^(test|example|sample|fake|domain|email|mail)\.(com|con|net|org|co)$", re.I)
TYPO_DOMAIN = re.compile(r"(gamail|gmial|gnail|gmal|gmaill|gmali|yahooo|hotmial|outlok|iclould)\.|\.(con|cmo|comm|ocm|vom)$", re.I)
# Role words match with separator suffixes too: guestlist.service@ was the
# real incident address and the old exact-match regex let it through.
ROLE_LOCAL = re.compile(
    r"^(info|contact|support|admin|sales|team|office|events?|tickets?|booking|press|marketing|billing"
    r"|newsletter|service|guestlist|help|jobs|careers|hr|hello|mail|enquiries|inquiries|boxoffice)([._+-].*)?$",
    re.I,
)
SERVICE_DOMAINS = {"eventbrite.com"}  # platform staff/relay, never a fan

FAR_STATES = {"CA", "WA", "NV", "FL", "TX", "IL", "MA", "IN", "AR", "GA", "NP",
              "OR", "AZ", "CO", "MN", "OH", "MI", "IA"}
STATE_NAMES = {  # full names typed by buyers -> code (region check needs codes)
    "california": "CA", "washington": "WA", "nevada": "NV", "florida": "FL",
    "texas": "TX", "illinois": "IL", "massachusetts": "MA", "indiana": "IN",
    "arkansas": "AR", "georgia": "GA", "oregon": "OR", "arizona": "AZ",
    "colorado": "CO", "minnesota": "MN", "ohio": "OH", "michigan": "MI",
    "iowa": "IA", "new york": "NY", "new jersey": "NJ", "connecticut": "CT",
    "pennsylvania": "PA",
}
# Superset of the old list, merged from score_unemailed.py so a city-only row
# ("Santa Clara", blank state) cannot pass (Codex audit / Devang incident).
FAR_CITIES = {
    "san francisco", "los angeles", "chicago", "houston", "seattle", "austin",
    "san diego", "las vegas", "oakland", "berkeley", "san jose", "sacramento",
    "boston", "miami", "atlanta", "des moines", "arcadia", "ashburn",
    "bentonville", "bothell", "brookline", "burlingame", "concord",
    "daly city", "danville", "emeryville", "fayetteville", "fremont",
    "harbor city", "harringay", "hayward", "indianapolis", "inglewood",
    "irvine", "los altos", "milpitas", "modesto", "moraga", "mountain view",
    "mumbai", "nassau", "north hollywood", "palo alto", "perris",
    "pleasanton", "randolph", "riverside", "san bruno", "san leandro",
    "san mateo", "san ramon", "santa clara", "sarasota",
    "south san francisco", "sunnyvale", "torrance", "tracy", "vallejo",
    "worcester",
}
IN_REGION = {"NY", "NJ", "NYC", "NY-NJ METRO"}
TRUTHY = {"true", "yes", "1"}


def die(msg):
    sys.exit(f"GATE HARD ERROR: {msg}")


def norm_email(v):
    e = (v or "").strip().lower()
    return e


def sha256_file(p):
    return hashlib.sha256(Path(p).read_bytes()).hexdigest()


def open_csv(p):
    # utf-8-sig: Klaviyo and Eventbrite exports carry a BOM.
    return open(p, encoding="utf-8-sig", newline="")


def load_suppressions():
    """All three suppression sources are MANDATORY (Codex audit #14): a
    silently missing kill list was a live failure mode, never an option."""
    sup = {}
    sup_file = DATA_ROOT / "sender" / "suppressed.csv" if os.environ.get("GATE_TEST_ROOT") else HERE / "suppressed.csv"
    if not sup_file.exists():
        die(f"suppressed.csv missing at {sup_file} — the permanent kill list is required, always.")
    with open_csv(sup_file) as fh:
        rd = csv.DictReader(fh)
        if rd.fieldnames is None or [f.strip() for f in rd.fieldnames[:2]] != ["email", "reason"]:
            die(f"suppressed.csv header must be 'email,reason', got: {rd.fieldnames}")
        for i, r in enumerate(rd, start=2):
            e = norm_email(r.get("email"))
            if not e or "@" not in e:
                print(f"WARNING: suppressed.csv line {i} malformed, ignoring: {r}", file=sys.stderr)
                continue
            sup[e] = "suppressed.csv: " + (r.get("reason") or "")

    excl_matches = sorted((DATA_ROOT / "downloads-stuff").glob("Profile Exclusions*.csv"))
    if not excl_matches:
        die(f"Klaviyo Profile Exclusions CSV missing under {DATA_ROOT / 'downloads-stuff'} — required.")
    excl = excl_matches[-1]
    age_days = (time.time() - excl.stat().st_mtime) / 86400
    if age_days > 30:
        print(f"WARNING: exclusions snapshot {excl.name} is {age_days:.0f} days old — "
              f"re-export from Klaviyo if the account is still active.", file=sys.stderr)
    with open_csv(excl) as fh:
        for r in csv.DictReader(fh):
            e = norm_email(r.get("Email Address"))
            if e:
                sup.setdefault(e, "klaviyo exclusion: " + (r.get("Exclusion Reason") or ""))

    master = DATA_ROOT / "garam-masala-audience" / "Garam_Masala_Master_Audience.xlsx"
    if not master.exists():
        die(f"master workbook missing at {master} — its Emailable=no column is a required suppression source.")
    try:
        import openpyxl
    except ImportError:
        die("openpyxl unavailable — run with ../.venv/bin/python. NO BYPASS.")
    ws = openpyxl.load_workbook(master, read_only=True)["Master"]
    hdr = {c.value: i for i, c in enumerate(next(ws.iter_rows(min_row=1, max_row=1)))}
    for row in ws.iter_rows(min_row=2, values_only=True):
        e = norm_email(str(row[hdr["Email"]] or ""))
        if e and row[hdr["Emailable"]] == "no":
            sup.setdefault(e, "master: not emailable")
    return sup


def load_sent(campaign, all_campaigns):
    sent = {}
    log_dir = DATA_ROOT / "sender" if os.environ.get("GATE_TEST_ROOT") else HERE
    for lg in glob.glob(str(log_dir / "sent-log-*.json")):
        for k in json.load(open(lg)):
            camp, _, email = k.partition(":")
            if all_campaigns or camp == campaign:
                sent[email.lower()] = os.path.basename(lg)
    return sent


def load_moves():
    lu = (DATA_ROOT / "sender" / "location-updates.csv") if os.environ.get("GATE_TEST_ROOT") else HERE / "location-updates.csv"
    if not lu.exists():
        die(f"location-updates.csv missing at {lu} — create it with header 'email,new_region,new_city' "
            f"(empty below the header is fine). Movers must always be honored.")
    moved = {}
    with open_csv(lu) as fh:
        for r in csv.DictReader(fh):
            e = norm_email(r.get("email"))
            if e:
                moved[e] = (r.get("new_region") or "").strip().upper()
    return moved


def region_verdict(state_raw, city_raw, email, moved):
    """Return None to keep, or a drop reason. location-updates WINS over
    stale source data in BOTH directions (Codex audit #15): a mover INTO the
    region is kept even if their old typed city is far, and a mover OUT is
    dropped even if their old typed city was local."""
    if email in moved:
        return None if moved[email] in IN_REGION else f"moved away ({moved[email]})"
    st = (state_raw or "").strip().upper()
    st = STATE_NAMES.get(st.lower(), st)
    city = (city_raw or "").strip().lower()
    if st in FAR_STATES or city in FAR_CITIES:
        return f"far from NYC ({st or city})"
    return None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("src")
    ap.add_argument("dst")
    ap.add_argument("--campaign", required=True,
                    help="MUST match the sender's SHOW.campaign; the sender rejects mismatched manifests")
    ap.add_argument("--region", choices=["nyc", "any"], default="nyc")
    ap.add_argument("--exclude-list", action="append", default=[],
                    help="another queue CSV whose recipients must be excluded (parallel channels)")
    ap.add_argument("--exclude-all-campaigns", action="store_true",
                    help="subtract people emailed in ANY campaign, not just --campaign")
    args = ap.parse_args()
    src, dst = Path(args.src), Path(args.dst)
    if not src.exists():
        die(f"input not found: {src}")

    # Invalidate any previous manifest for this output FIRST: a failed build
    # must never leave a stale-but-valid signed queue behind (Codex audit #4).
    manifest_path = Path(f"{dst}.manifest.json")
    legacy_sig = Path(f"{dst}.sig")
    for stale in (manifest_path, legacy_sig):
        if stale.exists():
            stale.unlink()

    suppressed = load_suppressions()
    sent = load_sent(args.campaign, args.exclude_all_campaigns)
    moved = load_moves() if args.region == "nyc" else {}

    explicit_excl = set()
    for xl in args.exclude_list:
        with open_csv(xl) as fh:
            for r in csv.DictReader(fh):
                e = norm_email(r.get("email"))
                if e:
                    explicit_excl.add(e)

    # Pass 1 — aggregate rows per email BEFORE judging, so a disqualifying
    # state on any duplicate row (unsubscribed) wins over a clean earlier
    # row (Codex audit #13).
    people = {}
    order = []
    dropped = []
    with open_csv(src) as fh:
        rd = csv.DictReader(fh)
        if rd.fieldnames is None or "email" not in [f.strip() for f in rd.fieldnames]:
            die(f"input {src} has no 'email' column (header: {rd.fieldnames})")
        for r in rd:
            e = norm_email(r.get("email"))
            if not e:
                dropped.append({**r, "dropped_because": "blank email"})
                continue
            if e not in people:
                people[e] = dict(r)
                people[e]["email"] = e  # canonicalized (lowercase, trimmed)
                people[e]["_unsub"] = str(r.get("unsubscribed", "")).strip().lower() in TRUTHY
                order.append(e)
            else:
                merged = people[e]
                for k, v in r.items():
                    if v and not merged.get(k):
                        merged[k] = v
                if str(r.get("unsubscribed", "")).strip().lower() in TRUTHY:
                    merged["_unsub"] = True
                dropped.append({**r, "dropped_because": "duplicate in file (aggregated)"})

    kept = []
    for e in order:
        r = people[e]
        local, _, domain = e.partition("@")

        def drop(why, row=r):
            row_out = {k: v for k, v in row.items() if k != "_unsub"}
            row_out["dropped_because"] = why
            dropped.append(row_out)

        if not EMAIL_RE.match(e) or ".." in e:
            drop("invalid syntax"); continue
        if JUNK_LOCAL.match(local) or JUNK_DOMAIN.match(domain):
            drop("junk/test address"); continue
        if TYPO_DOMAIN.search(domain):
            drop("typo domain (will bounce)"); continue
        if ROLE_LOCAL.match(local) and not domain.endswith("garammasaladating.com"):
            drop("role/service address, not a person"); continue
        if domain in SERVICE_DOMAINS:
            drop("platform service domain, not a person"); continue
        if e in suppressed:
            drop(suppressed[e]); continue
        if e in sent:
            drop(f"already sent ({sent[e]})"); continue
        if e in explicit_excl:
            drop("in --exclude-list queue"); continue
        if r["_unsub"]:
            drop("unsubscribed flag"); continue
        if args.region == "nyc":
            why = region_verdict(r.get("state"), r.get("city"), e, moved)
            if why:
                drop(why); continue
        del r["_unsub"]
        kept.append(r)

    if not kept:
        die("ZERO rows survived — nothing written, and any previous manifest for this output was invalidated.")

    # Parallel-channel safety net (Codex audit #17): automatic, not opt-in —
    # the output must not overlap any OTHER manifest-signed queue of the same
    # campaign sitting in the same directory.
    kept_set = {r["email"] for r in kept}
    for other_manifest in dst.parent.glob("*.manifest.json"):
        if other_manifest == manifest_path:
            continue
        try:
            om = json.load(open(other_manifest))
        except (json.JSONDecodeError, OSError):
            continue
        if om.get("campaign") != args.campaign:
            continue
        other_csv = Path(str(other_manifest)[: -len(".manifest.json")])
        if not other_csv.exists():
            continue
        with open_csv(other_csv) as fh:
            overlap = kept_set & {norm_email(r.get("email")) for r in csv.DictReader(fh)}
        if overlap:
            die(f"{len(overlap)} recipients overlap already-signed queue {other_csv.name} "
                f"(same campaign). Pass it via --exclude-list and rebuild.")

    # Atomic publish: temps in the same dir, manifest LAST so a valid
    # manifest implies every artifact is complete (Codex audit #4).
    fieldnames = [k for k in kept[0].keys()]
    tmp_q = Path(f"{dst}.tmp")
    with open(tmp_q, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        w.writeheader(); w.writerows(kept)
    tmp_d = Path(f"{dst}.dropped.csv.tmp")
    with open(tmp_d, "w", newline="") as f:
        if dropped:
            d_fields = sorted({k for d in dropped for k in d}, key=lambda k: (k == "dropped_because", k))
            w = csv.DictWriter(f, fieldnames=d_fields, extrasaction="ignore")
            w.writeheader(); w.writerows(dropped)
    os.replace(tmp_q, dst)
    os.replace(tmp_d, f"{dst}.dropped.csv")

    manifest = {
        "gate_version": GATE_VERSION,
        "sha256": sha256_file(dst),
        "campaign": args.campaign,
        "region": args.region,
        "output": dst.name,
        "built_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "kept": len(kept),
        "dropped": len(dropped),
        "source_sha256": sha256_file(src),
    }
    tmp_m = Path(f"{manifest_path}.tmp")
    tmp_m.write_text(json.dumps(manifest, indent=2) + "\n")
    os.replace(tmp_m, manifest_path)

    reasons = {}
    for d in dropped:
        key = d["dropped_because"].split(":")[0]
        reasons[key] = reasons.get(key, 0) + 1
    print(f"GATE PASSED: kept {len(kept)}, dropped {len(dropped)} -> {dst} (+.manifest.json, +.dropped.csv audit)")
    print(f"  campaign={args.campaign} region={args.region} gate_version={GATE_VERSION}")
    for why, n in sorted(reasons.items(), key=lambda t: -t[1]):
        print(f"  dropped {n}: {why}")


if __name__ == "__main__":
    main()
